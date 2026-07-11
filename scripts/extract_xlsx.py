#!/usr/bin/env python3
"""Extract Digilex_Financial_Control_Center.xlsx into clean JSON files under src/data/.

Usage: python3 scripts/extract_xlsx.py [path/to/workbook.xlsx]
Re-run this whenever the source workbook is updated, then restart the app
(or use Settings > Reset to original data) to reseed with the new figures.
"""
import json
import re
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl

SRC = sys.argv[1] if len(sys.argv) > 1 else "Digilex_Financial_Control_Center.xlsx"
OUT = Path(__file__).resolve().parent.parent / "src" / "data"
OUT.mkdir(parents=True, exist_ok=True)

wb = openpyxl.load_workbook(SRC, data_only=True)


def d(v):
    """Serialize a cell value: dates -> ISO string, everything else passthrough."""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d") if not isinstance(v, datetime) or (
            v.hour == 0 and v.minute == 0 and v.second == 0
        ) else v.isoformat()
    return v


def row_vals(ws, r, max_col):
    return [d(ws.cell(row=r, column=c).value) for c in range(1, max_col + 1)]


def is_blank(vals):
    return all(v is None or v == "" for v in vals)


def table(ws, header_row, max_col, start_row=None, end_row=None):
    headers = [h if h else f"col{i}" for i, h in enumerate(row_vals(ws, header_row, max_col))]
    start_row = start_row or header_row + 1
    end_row = end_row or ws.max_row
    rows = []
    for r in range(start_row, end_row + 1):
        vals = row_vals(ws, r, max_col)
        if is_blank(vals):
            continue
        # Skip in-sheet "TOTAL" / "TOTALS" summary rows (no primary key, a
        # label cell says TOTAL...) so they aren't double-counted as data.
        if vals[0] is None and any(
            isinstance(v, str) and v.strip().upper().startswith("TOTAL") for v in vals
        ):
            continue
        rows.append({headers[i]: vals[i] for i in range(len(headers))})
    return headers, rows


def dump(name, obj):
    path = OUT / f"{name}.json"
    path.write_text(json.dumps(obj, ensure_ascii=False, default=str))
    print(f"wrote {path} ({path.stat().st_size / 1024:.1f} KB)")


# ---------------------------------------------------------------------------
# 1. DASHBOARD
# ---------------------------------------------------------------------------
ws = wb["Dashboard"]
kpi_labels_1 = row_vals(ws, 4, 5)
kpi_values_1 = row_vals(ws, 5, 5)
kpi_labels_2 = row_vals(ws, 6, 5)
kpi_values_2 = row_vals(ws, 7, 5)
kpi_labels_3 = row_vals(ws, 8, 5)
kpi_values_3 = row_vals(ws, 9, 5)
kpi_labels_4 = row_vals(ws, 10, 3)
kpi_values_4 = row_vals(ws, 11, 3)

kpis = {}
for labels, values in [
    (kpi_labels_1, kpi_values_1),
    (kpi_labels_2, kpi_values_2),
    (kpi_labels_3, kpi_values_3),
    (kpi_labels_4, kpi_values_4),
]:
    for label, value in zip(labels, values):
        if label:
            kpis[label] = value

notes = []
for r in range(13, ws.max_row + 1):
    v = ws.cell(row=r, column=1).value
    if v and isinstance(v, str) and len(v) > 20:
        notes.append(v)

dump(
    "dashboard",
    {
        "title": ws.cell(row=1, column=1).value,
        "subtitle": ws.cell(row=2, column=1).value,
        "selectedMonth": d(ws.cell(row=3, column=3).value),
        "kpis": kpis,
        "notes": notes,
    },
)

# ---------------------------------------------------------------------------
# 2. INCOME TRACKER
# ---------------------------------------------------------------------------
ws = wb["Income Tracker"]
quick_totals = {
    "dailyDate": d(ws.cell(row=4, column=2).value),
    "dailyIncome": ws.cell(row=4, column=4).value,
    "weekStart": d(ws.cell(row=5, column=2).value),
    "weeklyIncome": ws.cell(row=5, column=4).value,
    "month": d(ws.cell(row=6, column=2).value),
    "monthlyIncome": ws.cell(row=6, column=4).value,
    "year": ws.cell(row=7, column=2).value,
    "yearlyIncome": ws.cell(row=7, column=4).value,
}
headers, rows = table(ws, 9, 13)
dump("incomeTracker", {"quickTotals": quick_totals, "columns": headers, "rows": rows})

# ---------------------------------------------------------------------------
# 3. EXPENSE TRACKER
# ---------------------------------------------------------------------------
ws = wb["Expense Tracker"]
note = ws.cell(row=2, column=1).value
headers, rows = table(ws, 4, 8)
category_totals = []
for r in range(5, ws.max_row + 1):
    label = ws.cell(row=r, column=10).value
    val = ws.cell(row=r, column=11).value
    if label:
        category_totals.append({"category": label, "total": val})
dump(
    "expenseTracker",
    {"note": note, "columns": headers, "rows": rows, "categoryTotals": category_totals},
)

# ---------------------------------------------------------------------------
# 4. FACEBOOK ADS TRACKER
# ---------------------------------------------------------------------------
ws = wb["Facebook Ads Tracker"]
note = ws.cell(row=2, column=1).value
acc_headers, acc_rows = table(ws, 5, 6, start_row=6, end_row=12)
total_confirmed = ws.cell(row=13, column=3).value
txn_headers, txn_rows = table(ws, 16, 9)
dump(
    "facebookAds",
    {
        "note": note,
        "accountCoverage": acc_rows,
        "totalConfirmedSpend": total_confirmed,
        "transactions": txn_rows,
    },
)

# ---------------------------------------------------------------------------
# 5. CREDIT CARD RECONCILIATION
# ---------------------------------------------------------------------------
ws = wb["Credit Card Reconciliation"]
note = ws.cell(row=2, column=1).value
headers, rows = table(ws, 4, 8)
dump("creditCardReconciliation", {"note": note, "columns": headers, "rows": rows})

# ---------------------------------------------------------------------------
# 6. CASH FLOW
# ---------------------------------------------------------------------------
ws = wb["Cash Flow"]
note = ws.cell(row=2, column=1).value
headers, rows = table(ws, 4, 6)
dump("cashFlow", {"note": note, "columns": headers, "rows": rows})

# ---------------------------------------------------------------------------
# 7. BILLS & PAYMENT REMINDERS  (card-block layout)
# ---------------------------------------------------------------------------
ws = wb["Bills & Payment Reminders"]
note = ws.cell(row=2, column=1).value
months = []
current_month = None
current_bill = None
r = 4
while r <= ws.max_row:
    b_val = ws.cell(row=r, column=2).value
    a_val = ws.cell(row=r, column=1).value
    e_val = ws.cell(row=r, column=5).value
    f_val = ws.cell(row=r, column=6).value
    if a_val and not b_val and re.match(r"^[A-Za-z]+ \d{4}$", str(a_val)):
        current_month = {"month": a_val, "bills": []}
        months.append(current_month)
        r += 1
        continue
    if b_val and e_val and str(e_val).lower().startswith(("due in", "confirm due day")):
        current_bill = {
            "name": b_val,
            "dueInLabel": e_val,
            "dueDate": None,
            "totalAmountDue": None,
            "category": None,
            "paymentAccount": None,
            "status": None,
            "autoDebit": None,
            "notes": [],
        }
        if current_month is not None:
            current_month["bills"].append(current_bill)
        r += 1
        continue
    if current_bill is not None and b_val:
        key = str(b_val).strip().lower()
        if key == "due date":
            current_bill["dueDate"] = d(ws.cell(row=r, column=3).value)
            current_bill["totalAmountDue"] = ws.cell(row=r, column=6).value
        elif key == "category":
            current_bill["category"] = ws.cell(row=r, column=3).value
            current_bill["paymentAccount"] = ws.cell(row=r, column=6).value
        elif key == "status":
            current_bill["status"] = ws.cell(row=r, column=3).value
            current_bill["autoDebit"] = ws.cell(row=r, column=6).value
        elif key == "total amount due":
            current_bill["totalAmountDue"] = ws.cell(row=r, column=3).value
            current_bill["paymentAccount"] = ws.cell(row=r, column=6).value
        elif isinstance(b_val, str) and len(b_val) > 15:
            current_bill["notes"].append(b_val)
    elif current_bill is not None and a_val is None and b_val is None and not is_blank(
        row_vals(ws, r, 6)
    ):
        pass
    r += 1
dump("billsReminders", {"note": note, "months": months})

# ---------------------------------------------------------------------------
# 8. MONTHLY P&L
# ---------------------------------------------------------------------------
ws = wb["Monthly P&L"]
note = ws.cell(row=2, column=1).value
headers, rows = table(ws, 3, 5)
dump("monthlyPL", {"note": note, "columns": headers, "rows": rows})

# ---------------------------------------------------------------------------
# 9. MONTHLY BOOKKEEPING (month sections -> INCOME / EXPENSES blocks)
# ---------------------------------------------------------------------------
ws = wb["Monthly Bookkeeping"]
note = ws.cell(row=2, column=1).value
months = []
r = 4
current = None
section = None
while r <= ws.max_row:
    vals = row_vals(ws, r, 6)
    a = vals[0]
    if a and re.match(r"^[A-Za-z]+ \d{4}$", str(a)) and all(v is None for v in vals[1:]):
        current = {"month": a, "income": [], "incomeTotal": None, "expenses": [], "expenseTotal": None, "net": None}
        months.append(current)
        section = None
        r += 1
        continue
    if a in ("INCOME", "EXPENSES") and all(v is None for v in vals[1:]):
        section = a
        r += 1
        continue
    if a == "Date" and vals[1] == "Type":
        r += 1
        continue
    if current is not None and vals[3] == "TOTAL INCOME (excl. Owner Transfer)":
        current["incomeTotal"] = vals[4]
        r += 1
        continue
    if current is not None and vals[3] == "TOTAL EXPENSES (excl. Owner Transfer / Personal)":
        current["expenseTotal"] = vals[4]
        r += 1
        continue
    if current is not None and str(vals[3]).startswith("NET -") if vals[3] else False:
        current["net"] = vals[4]
        r += 1
        continue
    if current is not None and section and a is not None:
        entry = {
            "date": vals[0] if not (isinstance(vals[0], str) and re.match(r"^[A-Za-z]+ \d{4}$", vals[0])) else None,
            "type": vals[1],
            "category": vals[2],
            "description": vals[3],
            "amount": vals[4],
            "notes": vals[5],
        }
        if section == "INCOME":
            current["income"].append(entry)
        else:
            current["expenses"].append(entry)
    r += 1
dump("monthlyBookkeeping", {"note": note, "months": months})

# ---------------------------------------------------------------------------
# 10. ORDERS DATABASE (big)
# ---------------------------------------------------------------------------
ws = wb["Orders Database"]
headers, rows = table(ws, 3, 19)
dump("ordersDatabase", {"columns": headers, "rows": rows})

# ---------------------------------------------------------------------------
# 11. FULFILLMENT SOA RECONCILIATION
# ---------------------------------------------------------------------------
ws = wb["Fulfillment SOA Reconciliation"]
warning = ws.cell(row=3, column=1).value
headers, rows = table(ws, 4, 16)
dump("fulfillmentSOAReconciliation", {"warning": warning, "columns": headers, "rows": rows})

# ---------------------------------------------------------------------------
# 12. SOA BREAKDOWN & VERIFICATION
# ---------------------------------------------------------------------------
ws = wb["SOA Breakdown & Verification"]
dup_check = {
    "label": ws.cell(row=4, column=1).value,
    "count": ws.cell(row=4, column=9).value,
    "message": ws.cell(row=4, column=11).value,
}
explanation_title = ws.cell(row=6, column=1).value
explanation_text = ws.cell(row=7, column=1).value
batch_headers, batch_rows = table(ws, 10, 6, start_row=11, end_row=22)
verif_headers, verif_rows = table(ws, 27, 19, start_row=28, end_row=39)
verif_summary = ws.cell(row=40, column=1).value
shortfall_note1 = ws.cell(row=42, column=1).value
shortfall_note2 = ws.cell(row=43, column=1).value
shortfall_headers, shortfall_rows = table(ws, 44, 5, start_row=45, end_row=47)
dump(
    "soaBreakdown",
    {
        "duplicateCheck": dup_check,
        "explanationTitle": explanation_title,
        "explanationText": explanation_text,
        "batchTable": batch_rows,
        "verificationTable": verif_rows,
        "verificationSummary": verif_summary,
        "shortfallNotes": [shortfall_note1, shortfall_note2],
        "shortfallTable": shortfall_rows,
    },
)

# ---------------------------------------------------------------------------
# 13. FULFILLMENT VERIFICATION
# ---------------------------------------------------------------------------
ws = wb["Fulfillment Verification"]
note = ws.cell(row=2, column=1).value
summary = {
    "label": ws.cell(row=5, column=1).value,
    "totalFeeParcels": ws.cell(row=5, column=5).value,
    "matchedLabel": ws.cell(row=6, column=1).value,
    "matched": ws.cell(row=6, column=5).value,
    "unmatchedLabel": ws.cell(row=7, column=1).value,
    "unmatchedCount": ws.cell(row=7, column=5).value,
    "unmatchedFee": ws.cell(row=7, column=6).value,
}
caveat = ws.cell(row=9, column=1).value
headers, rows = table(ws, 12, 7)
dump(
    "fulfillmentVerification",
    {"note": note, "summary": summary, "caveat": caveat, "columns": headers, "rows": rows},
)

# ---------------------------------------------------------------------------
# 14. POS ORDER RECONCILIATION
# ---------------------------------------------------------------------------
ws = wb["POS Order Reconciliation"]
note = ws.cell(row=2, column=1).value
summary_rows = []
for r in range(4, 10):
    label = ws.cell(row=r, column=1).value
    if label:
        summary_rows.append(
            {"label": label, "count": ws.cell(row=r, column=5).value, "value": ws.cell(row=r, column=6).value}
        )
rts = {
    "deliveredLabel": ws.cell(row=12, column=1).value,
    "delivered": ws.cell(row=12, column=7).value,
    "rtsLabel": ws.cell(row=13, column=1).value,
    "rts": ws.cell(row=13, column=7).value,
    "transitLabel": ws.cell(row=14, column=1).value,
    "transit": ws.cell(row=14, column=7).value,
    "rateLabel": ws.cell(row=15, column=1).value,
    "rate": ws.cell(row=15, column=7).value,
}
monthly_headers, monthly_rows = table(ws, 19, 5, start_row=20, end_row=23)
coverage = ws.cell(row=24, column=1).value
coverage_val = ws.cell(row=24, column=4).value
headers, rows = table(ws, 25, 10)
dump(
    "posOrderReconciliation",
    {
        "note": note,
        "summary": summary_rows,
        "rts": rts,
        "monthlyRts": monthly_rows,
        "coverage": {"label": coverage, "value": coverage_val},
        "columns": headers,
        "rows": rows,
    },
)

# ---------------------------------------------------------------------------
# 15. EVIDENCE - DELIVERED NOT IN SOA
# ---------------------------------------------------------------------------
ws = wb["Evidence - Delivered Not in SOA"]
note = ws.cell(row=2, column=1).value
summary_rows = []
for r in range(5, 11):
    label = ws.cell(row=r, column=1).value
    if label:
        summary_rows.append(
            {"label": label, "count": ws.cell(row=r, column=5).value, "value": ws.cell(row=r, column=6).value}
        )
headers, rows = table(ws, 12, 12)
dump("evidenceNotInSOA", {"note": note, "summary": summary_rows, "columns": headers, "rows": rows})

# ---------------------------------------------------------------------------
# 16. FOR DIGILEX - FOLLOW-UP LIST
# ---------------------------------------------------------------------------
ws = wb["For Digilex - Follow-Up List"]
meta = {
    "preparedBy": ws.cell(row=2, column=2).value,
    "datePrepared": d(ws.cell(row=3, column=2).value),
    "purpose": ws.cell(row=4, column=2).value,
    "totalOrders": ws.cell(row=6, column=2).value,
    "totalValue": ws.cell(row=6, column=4).value,
}
headers, rows = table(ws, 8, 8)
dump("followUpList", {"meta": meta, "columns": headers, "rows": rows})

# ---------------------------------------------------------------------------
# 17. SETTINGS
# ---------------------------------------------------------------------------
ws = wb["Settings"]
business = {
    "businessName": ws.cell(row=4, column=2).value,
    "currency": ws.cell(row=5, column=2).value,
    "fulfillmentPartner": ws.cell(row=6, column=2).value,
    "reportGenerated": d(ws.cell(row=7, column=2).value),
}
prod_headers, products = table(ws, 10, 6, start_row=11, end_row=14)
fee_benchmarks = []
for r in range(17, 22):
    label = ws.cell(row=r, column=1).value
    if label:
        fee_benchmarks.append(
            {"label": label, "value": ws.cell(row=r, column=2).value, "note": ws.cell(row=r, column=3).value}
        )
fixed_headers, fixed_expenses = table(ws, 25, 2, start_row=26, end_row=30)
total_fixed = ws.cell(row=31, column=2).value
tax = ws.cell(row=33, column=2).value
dropdown_headers, dropdown_rows = table(ws, 35, 4, start_row=36, end_row=53)
dropdowns = {h: [] for h in dropdown_headers}
for row in dropdown_rows:
    for h in dropdown_headers:
        if row.get(h):
            dropdowns[h].append(row[h])
dump(
    "settings",
    {
        "business": business,
        "products": products,
        "feeBenchmarks": fee_benchmarks,
        "fixedExpenses": fixed_expenses,
        "totalFixedExpenses": total_fixed,
        "tax": tax,
        "dropdowns": dropdowns,
    },
)

print("\nDone.")
